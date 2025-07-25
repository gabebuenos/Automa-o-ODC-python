import openpyxl 
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep



planilha_clientes = openpyxl.load_workbook('dados_cpf.xlsx')
pagina_clientes = planilha_clientes['Planilha1']

driver = webdriver.Chrome()  
driver.get('https://gabriel-bueno.outsystemscloud.com/ConsultaCpf/')

for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
    Nome,cpf,valor, vencimento = linha
    sleep(8)
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='Input_CPF']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-primary ThemeGrid_MarginGutter']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
    status = driver.find_element(By.XPATH,"//span[@class='oi ThemeGrid_MarginGutter']")
    if status.text == 'Em Dia':
        data_pagamento = driver.find_element(By.XPATH,"//span[@class='ThemeGrid_MarginGutter']")
        metodo_pagamento = driver.find_element(By.XPATH,"//span[@style='font-weight: bold;']")
        data_pagamento_limpo = data_pagamento.text
        metodo_pagamento_limpo = metodo_pagamento.text
        planilha_fechamento = openpyxl.load_workbook('Planilha_fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Planilha1']

        pagina_fechamento.append([Nome,valor,cpf,vencimento,'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        planilha_fechamento.save('Planilha_fechamento.xlsx')
    else:
        planilha_fechamento = openpyxl.load_workbook('Planilha_fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Planilha1']

        pagina_fechamento.append([Nome,valor,cpf,vencimento,'pendente'])
        planilha_fechamento.save('Planilha_fechamento.xlsx')